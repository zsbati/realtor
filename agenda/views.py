from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django import forms
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.utils import timezone
from django.contrib.auth.views import PasswordResetView
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import Visit
from .forms import VisitFilterForm, VisitForm, CustomUserCreationForm, CustomUserChangeForm, PasswordChangeForm
import datetime

# Create your views here.

@login_required
def dashboard(request):
    today = datetime.date.today()
    
    # Get today's visits
    today_visits = Visit.objects.filter(
        scheduled_date__date=today
    ).order_by('scheduled_date')
    
    # Get upcoming visits (next 7 days)
    upcoming_visits = Visit.objects.filter(
        scheduled_date__date__gte=today,
        scheduled_date__date__lte=today + datetime.timedelta(days=7)
    ).exclude(
        scheduled_date__date=today
    ).order_by('scheduled_date')[:5]
    
    # Get recent visits (last 7 days)
    recent_visits = Visit.objects.filter(
        scheduled_date__date__gte=today - datetime.timedelta(days=7),
        scheduled_date__date__lt=today
    ).order_by('-scheduled_date')[:5]
    
    # Get visit statistics
    total_visits = Visit.objects.count()
    today_visits_count = today_visits.count()
    upcoming_visits_count = upcoming_visits.count()
    
    # Get users (only for superusers)
    users = User.objects.all() if request.user.is_superuser else None
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        visits = []
        visits.extend(today_visits)
        visits.extend(upcoming_visits)
        visits.extend(recent_visits)
        return export_visits_to_excel(visits, f'visitas_dashboard_{timezone.now().strftime("%Y%m%d")}.xlsx')

    context = {
        'today_visits': today_visits,
        'upcoming_visits': upcoming_visits,
        'recent_visits': recent_visits,
        'total_visits': total_visits,
        'today_visits_count': today_visits_count,
        'upcoming_visits_count': upcoming_visits_count,
        'users': users,
        'user_list_url': 'users:user_list',
    }
    return render(request, 'dashboard.html', context)

@login_required
def reports(request):
    """View to display past visits with filtering capabilities."""
    today = datetime.date.today()
    
    if request.method == 'POST':
        form = VisitFilterForm(request.POST)
        if form.is_valid():
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            visit_type = form.cleaned_data['visit_type']
            status = form.cleaned_data['status']
            
            # Filter for completed visits by default
            visits = Visit.objects.filter(status='completed')
            
            if start_date:
                visits = visits.filter(scheduled_date__gte=start_date)
            if end_date:
                visits = visits.filter(scheduled_date__lte=end_date)
            if visit_type:
                visits = visits.filter(visit_type=visit_type)
            
            visits = visits.order_by('-scheduled_date')
            
            if request.GET.get('export') == 'excel':
                return export_visits_to_excel(visits)
            
            context = {
                'form': form,
                'visits': visits,
            }
            return render(request, 'agenda/reports.html', context)
    else:
        form = VisitFilterForm()
        # Filter for completed visits by default
        visits = Visit.objects.filter(status='completed').order_by('-scheduled_date')
        
        if request.GET.get('export') == 'excel':
            return export_visits_to_excel(visits)
        
        context = {
            'form': form,
            'visits': visits,
        }
        return render(request, 'agenda/reports.html', context)

@login_required
def visit_list(request):
    today = datetime.date.today()
    visits = Visit.objects.filter(scheduled_date__date=today).order_by('scheduled_date')
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        return export_visits_to_excel(visits, f'visitas_{timezone.now().strftime("%Y%m%d")}.xlsx')

    context = {
        'visits': visits,
    }
    return render(request, 'visit_list.html', context)

@login_required
def visit_detail(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    return render(request, 'visit_detail.html', {'visit': visit})

@login_required
def password_change(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password1']
            request.user.set_password(new_password)
            request.user.save()
            messages.success(request, 'Senha alterada com sucesso.')
            if request.user.is_superuser:
                return redirect('users:user_list')
            return redirect('agenda:visit_list')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'password_change.html', {'form': form})

@login_required
def logout(request):
    """Custom logout view that handles GET requests."""
    auth_logout(request)
    return redirect('agenda:login')

@login_required
def visit_create(request):
    if request.method == 'POST':
        form = VisitForm(request.POST)
        if form.is_valid():
            visit = form.save(commit=False)
            visit.created_by = request.user
            visit.save()
            return redirect('agenda:visit_list')
    else:
        form = VisitForm()
    return render(request, 'visit_form.html', {'form': form})

@login_required
def visit_edit(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    if request.method == 'POST':
        form = VisitForm(request.POST, instance=visit)
        if form.is_valid():
            visit = form.save()
            return redirect('agenda:visit_list')
    else:
        form = VisitForm(instance=visit)
    return render(request, 'visit_form.html', {'form': form})

@login_required
def visit_delete(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    if request.method == 'POST':
        visit.delete()
        return redirect('agenda:visit_list')
    return render(request, 'visit_confirm_delete.html', {'visit': visit})

@login_required
def export_visits_to_excel(visits, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitas"
    
    # Add headers
    headers = [
        'Data/Hora', 'Tipo', 'Nome', 'Transação', 'Morada', 'Email',
        'Telefone', 'Preço €', 'Estado', 'Comentários', 'Criado por',
        'Criado em', 'Atualizado em'
    ]
    
    # Style header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Add data
    for row_num, visit in enumerate(visits, 2):
        ws.cell(row=row_num, column=1, value=visit.scheduled_date.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row_num, column=2, value=visit.get_visit_type_display())
        ws.cell(row=row_num, column=3, value=visit.name)
        ws.cell(row=row_num, column=4, value=visit.get_transaction_type_display())
        ws.cell(row=row_num, column=5, value=visit.address)
        ws.cell(row=row_num, column=6, value=visit.email)
        ws.cell(row=row_num, column=7, value=visit.phone)
        ws.cell(row=row_num, column=8, value=str(visit.price))
        ws.cell(row=row_num, column=9, value=visit.get_status_display())
        ws.cell(row=row_num, column=10, value=visit.comment)
        ws.cell(row=row_num, column=11, value=visit.created_by.username)
        ws.cell(row=row_num, column=12, value=visit.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row_num, column=13, value=visit.updated_at.strftime("%d/%m/%Y %H:%M"))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def visit_list(request):
    today = datetime.date.today()
    visits = Visit.objects.filter(scheduled_date__date=today).order_by('scheduled_date')
    return render(request, 'visit_list.html', {'visits': visits})

@login_required
def visit_detail(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    return render(request, 'visit_detail.html', {'visit': visit})

@login_required
def password_change(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password1']
            request.user.set_password(new_password)
            request.user.save()
            messages.success(request, 'Senha alterada com sucesso.')
            if request.user.is_superuser:
                return redirect('users:user_list')
            return redirect('agenda:visit_list')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'password_change.html', {'form': form})

@login_required
def logout(request):
    """Custom logout view that handles GET requests."""
    auth_logout(request)
    return redirect('agenda:login')

@login_required
def visit_create(request):
    if request.method == 'POST':
        form = VisitForm(request.POST)
        if form.is_valid():
            visit = form.save(commit=False)
            visit.created_by = request.user
            visit.save()
            return redirect('agenda:visit_list')
    else:
        form = VisitForm()
    return render(request, 'visit_form.html', {'form': form})

@login_required
def visit_edit(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    if request.method == 'POST':
        form = VisitForm(request.POST, instance=visit)
        if form.is_valid():
            visit = form.save()
            return redirect('agenda:visit_list')
    else:
        form = VisitForm(instance=visit)
    return render(request, 'visit_form.html', {'form': form})

@login_required
def visit_delete(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    if request.method == 'POST':
        visit.delete()
        return redirect('agenda:visit_list')
    return render(request, 'visit_confirm_delete.html', {'visit': visit})

# User management views

def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('agenda:dashboard')
        else:
            messages.error(request, 'Nome de usuário ou senha inválidos.')
    return render(request, 'agenda/login.html')

def user_logout(request):
    logout(request)
    return redirect('agenda:login')

class CustomPasswordResetView(PasswordResetView):
    template_name = 'password_reset_form.html'
    email_template_name = 'password_reset_email.html'
    subject_template_name = 'password_reset_subject.txt'
    success_url = '/login/'
