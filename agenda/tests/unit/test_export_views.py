from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from django.utils import timezone
from agenda.models import Visit
from io import BytesIO
import openpyxl

User = get_user_model()

class ExportVisitsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(
            username='testuser',
            password='testpass123',
            is_staff=True
        )
        self.client.login(username='testuser', password='testpass123')
        
        # Create test visits with all required fields
        now = timezone.now()
        Visit.objects.create(
            title='Morning Visit',
            name='Client One',
            address='123 Test St',
            description='Test description',
            visit_type='visit',
            scheduled_date=now,
            status='scheduled',
            email='test@example.com',
            phone='123456789',
            comment='Test comment',
            price='100000',
            transaction_type='sale',
            created_by=self.user
        )
        Visit.objects.create(
            title='Afternoon Visit',
            name='Client Two',
            address='456 Test Ave',
            description='Another test',
            visit_type='meeting',
            scheduled_date=now,
            status='completed',
            email='test2@example.com',
            phone='987654321',
            comment='Test comment 2',
            price='200000',
            transaction_type='rent',
            created_by=self.user
        )
    
    def test_export_visits(self):
        # Test the export functionality
        response = self.client.get(reverse('agenda:export_visits'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertEqual(
            response['Content-Disposition'],
            'attachment; filename=my_visits_export.xlsx'
        )
        
        # Test the Excel file content
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = ['Title', 'Name', 'Scheduled Date', 'Status', 'Type', 'Address', 'Email', 'Phone', 'Notes']
        for header in expected_headers:
            self.assertIn(header, headers)
        
        # Check data rows (2 visits + header row)
        self.assertEqual(ws.max_row, 3)
