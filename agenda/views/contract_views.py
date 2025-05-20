from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from agenda.models.contract import Contract
from agenda.forms import ContractForm

@login_required
def contract_list(request):
    contracts = Contract.objects.all().order_by('-created_at')
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        return export_contracts_to_excel(contracts)
    
    context = {
        'contracts': contracts,
    }
    return render(request, 'agenda/contract_list.html', context)

@login_required
def contract_detail(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    return render(request, 'agenda/contract_detail.html', {'contract': contract})

@login_required
def contract_create(request):
    if request.method == 'POST':
        form = ContractForm(request.POST)
        if form.is_valid():
            contract = form.save(commit=False)
            contract.created_by = request.user
            contract.save()
            messages.success(request, 'Contrato criado com sucesso!')
            return redirect('agenda:contract_list')
    else:
        form = ContractForm()
    return render(request, 'agenda/contract_form.html', {'form': form})

@login_required
def contract_edit(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            contract = form.save()
            messages.success(request, 'Contrato atualizado com sucesso!')
            return redirect('agenda:contract_list')
    else:
        form = ContractForm(instance=contract)
    return render(request, 'agenda/contract_form.html', {'form': form})

@login_required
def contract_delete(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        contract.delete()
        return redirect('agenda:contract_list')
    return render(request, 'agenda/contract_confirm_delete.html', {'contract': contract})

@login_required
def export_contracts_to_excel(contracts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"
    
    # Add headers
    headers = [
        'Título', 'Tipo', 'Nome do Cliente', 'Morada da Propriedade', 'Valor',
        'Status', 'Data de Assinatura', 'Criado por', 'Criado em', 'Atualizado em',
        'Comentários'
    ]
    
    # Style header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Add data
    for row_num, contract in enumerate(contracts, 2):
        ws.cell(row=row_num, column=1, value=contract.title)
        ws.cell(row=row_num, column=2, value=contract.get_contract_type_display())
        ws.cell(row=row_num, column=3, value=contract.client_name)
        ws.cell(row=row_num, column=4, value=contract.property_address)
        ws.cell(row=row_num, column=5, value=str(contract.value))
        ws.cell(row=row_num, column=6, value=contract.get_status_display())
        ws.cell(row=row_num, column=7, value=contract.signing_date.strftime("%d/%m/%Y") if contract.signing_date else '')
        ws.cell(row=row_num, column=8, value=getattr(contract.created_by, 'username', 'N/A'))
        ws.cell(row=row_num, column=9, value=contract.created_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row_num, column=10, value=contract.updated_at.strftime("%d/%m/%Y %H:%M"))
        ws.cell(row=row_num, column=11, value=contract.comment)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="contratos_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
