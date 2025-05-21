from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django import forms
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.views import PasswordResetView
from django.contrib import messages
from django.contrib.auth import logout as auth_logout, login, authenticate
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime as dt, timedelta
from django.utils import timezone
from .models import Visit
from .forms import VisitForm, VisitFilterForm, CustomUserCreationForm, CustomUserChangeForm, PasswordChangeForm
import json
import datetime

# Create your views here.

@login_required
def dashboard(request):
    # Get current time in the user's timezone
    now = timezone.localtime(timezone.now())
    today = now.date()
    
    # Handle success message
    if request.GET.get('success') == '1':
        messages.success(request, 'Sua senha foi alterada com sucesso!')
    
    # Get today's visits (timezone-aware)
    today_start = timezone.make_aware(
        datetime.datetime.combine(today, datetime.time.min),
        timezone.get_current_timezone()
    )
    today_end = timezone.make_aware(
        datetime.datetime.combine(today, datetime.time.max),
        timezone.get_current_timezone()
    )
    
    today_visits = Visit.objects.filter(
        scheduled_date__date=today
    ).order_by('scheduled_date')
    
    # Get upcoming visits (next 7 days, timezone-aware)
    upcoming_start = today + datetime.timedelta(days=1)
    upcoming_end = today + datetime.timedelta(days=7)
    upcoming_visits = Visit.objects.filter(
        scheduled_date__date__range=[upcoming_start, upcoming_end]
    ).order_by('scheduled_date')
    
    # Get recent visits (past 7 days, timezone-aware)
    recent_start = today - datetime.timedelta(days=7)
    recent_end = today - datetime.timedelta(days=1)
    recent_visits = Visit.objects.filter(
        scheduled_date__date__range=[recent_start, recent_end]
    ).order_by('-scheduled_date')
    
    # Handle Excel export
    export_type = request.GET.get('export')
    if export_type:
        if export_type == 'today':
            return export_visits_to_excel(request, today_visits, 'visitas_hoje.xlsx')
        elif export_type == 'upcoming':
            return export_visits_to_excel(request, upcoming_visits, 'proximas_visitas.xlsx')
        elif export_type == 'recent':
            return export_visits_to_excel(request, recent_visits, 'visitas_recentes.xlsx')
    
    # Get counts for statistics cards
    today_visits_count = today_visits.count()
    upcoming_visits_count = upcoming_visits.count()
    recent_visits_count = recent_visits.count()
    total_visits_count = Visit.objects.count()
    
    # Get all users
    from django.contrib.auth import get_user_model
    User = get_user_model()
    users = User.objects.all().order_by('username')
    
    context = {
        'today_visits': today_visits,
        'upcoming_visits': upcoming_visits,
        'recent_visits': recent_visits,
        'now': now,
        'today_visits_count': today_visits_count,
        'upcoming_visits_count': upcoming_visits_count,
        'recent_visits_count': recent_visits_count,
        'total_visits': total_visits_count,
        'users': users,
    }
    
    return render(request, 'agenda/dashboard/dashboard.html', context)

@login_required
def reports(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    
    # Build query
    visits = Visit.objects.all()
    
    # Apply filters
    if start_date:
        start_date = timezone.make_aware(datetime.datetime.strptime(start_date, '%Y-%m-%d'))
        visits = visits.filter(scheduled_date__gte=start_date)
    
    if end_date:
        end_date = timezone.make_aware(datetime.datetime.strptime(end_date, '%Y-%m-%d'))
        visits = visits.filter(scheduled_date__lte=end_date)
    
    if status:
        visits = visits.filter(status=status)
    
    # Order by scheduled_date descending by default
    visits = visits.order_by('-scheduled_date')
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        filename = f'visitas_relatorio_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return export_visits_to_excel(request, visits, filename)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(visits, 25)
    
    try:
        visits = paginator.page(page)
    except PageNotAnInteger:
        visits = paginator.page(1)
    except EmptyPage:
        visits = paginator.page(paginator.num_pages)
    
    context = {
        'visits': visits,
        'filter_form': VisitFilterForm(request.GET or None),
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
        'status': status,
    }
    
    return render(request, 'agenda/reports.html', context)

@login_required
def visit_list(request):
    visits = Visit.objects.all().order_by('-scheduled_date')
    
    # Apply filters based on query parameters
    if request.GET.get('today') == '1':
        today = timezone.localdate()
        visits = visits.filter(scheduled_date__date=today)
    elif request.GET.get('status') == 'scheduled':
        visits = visits.filter(status='scheduled')
    
    # Handle Excel export
    if request.GET.get('export') == 'excel':
        filename = f'visitas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return export_visits_to_excel(request, visits, filename)
    
    # Add pagination
    paginator = Paginator(visits, 25)  # Show 25 visits per page
    page = request.GET.get('page')
    try:
        visits = paginator.page(page)
    except PageNotAnInteger:
        visits = paginator.page(1)
    except EmptyPage:
        visits = paginator.page(paginator.num_pages)
    
    context = {
        'visits': visits,
        'today_filter': request.GET.get('today') == '1',
        'scheduled_filter': request.GET.get('status') == 'scheduled',
    }
    return render(request, 'agenda/visits/visit_list.html', context)

@login_required
def calendar_view(request):
    """Render the calendar view."""
    context = {
        'page_title': 'Calendário de Visitas',
    }
    return render(request, 'agenda/calendar.html', context)

@login_required
def visit_calendar_events(request):
    """API endpoint for calendar events."""
    from django.http import JsonResponse
    from django.urls import reverse
    from django.conf import settings
    
    def json_response(data, status=200):
        response = JsonResponse(data, safe=False, json_dumps_params={'ensure_ascii': False})
        response['Access-Control-Allow-Origin'] = settings.ALLOWED_HOSTS[0] if hasattr(settings, 'ALLOWED_HOSTS') and settings.ALLOWED_HOSTS else '*'
        response['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'X-Requested-With, Content-Type, X-CSRFToken'
        return response
    
    if request.method == 'OPTIONS':
        return json_response({})
    
    try:
        # Get and validate parameters
        start = request.GET.get('start')
        end = request.GET.get('end')
        
        if not start or not end:
            return json_response({"error": "Start and end dates are required"}, status=400)
        
        # Parse dates
        try:
            if 'T' in start:
                start_date = timezone.datetime.fromisoformat(start)
                end_date = timezone.datetime.fromisoformat(end)
            else:
                start_date = timezone.make_aware(datetime.datetime.strptime(start, '%Y-%m-%d'))
                end_date = timezone.make_aware(datetime.datetime.strptime(end, '%Y-%m-%d'))
                
            if timezone.is_naive(start_date):
                start_date = timezone.make_aware(start_date)
            if timezone.is_naive(end_date):
                end_date = timezone.make_aware(end_date)
                
        except (ValueError, TypeError):
            return json_response({"error": "Invalid date format"}, status=400)
        
        # Get visits in date range
        try:
            visits = Visit.objects.filter(
                scheduled_date__range=(start_date, end_date)
            ).select_related('created_by').order_by('scheduled_date')
            
            events = []
            for visit in visits:
                if not visit.scheduled_date:
                    continue
                    
                event = {
                    'id': visit.id,
                    'title': visit.title or 'No title',
                    'start': visit.scheduled_date.isoformat(),
                    'end': (visit.scheduled_date + datetime.timedelta(hours=1)).isoformat(),
                    'url': reverse('agenda:visit_detail', args=[visit.id]),
                    'color': get_visit_color(visit.visit_type),
                    'extendedProps': {
                        'visit_type': visit.get_visit_type_display() if hasattr(visit, 'get_visit_type_display') else 'Not specified',
                        'status': visit.status,
                        'status_display': visit.get_status_display() if hasattr(visit, 'get_status_display') else 'Unknown',
                        'created_by': visit.created_by.username if visit.created_by else 'System',
                    },
                }
                events.append(event)
            
            return json_response(events)
            
        except Exception as e:
            return json_response({"error": "Error fetching visits"}, status=500)
            
    except Exception as e:
        return json_response({"error": "An error occurred"}, status=500)

def get_visit_color(visit_type):
    """Return color based on visit type."""
    colors = {
        'viewing': '#007bff',  # Blue
        'evaluation': '#28a745',  # Green
        'signature': '#17a2b8',  # Cyan
        'other': '#6c757d',  # Gray
    }
    return colors.get(visit_type, '#6c757d')  # Default to gray

@login_required
def visit_detail(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    return render(request, 'agenda/visits/detail.html', {'visit': visit})

@login_required
def visit_create(request):
    if request.method == 'POST':
        form = VisitForm(request.POST)
        if form.is_valid():
            try:
                visit = form.save(commit=False)
                visit.created_by = request.user
                
                # Ensure the datetime is timezone-aware
                if visit.scheduled_date:
                    if not timezone.is_aware(visit.scheduled_date):
                        visit.scheduled_date = timezone.make_aware(visit.scheduled_date)
                    # Ensure the timezone is set to the current timezone
                    visit.scheduled_date = timezone.localtime(visit.scheduled_date)
                
                visit.save()
                # Clear all existing messages
                storage = messages.get_messages(request)
                storage.used = True
                # Add success message
                messages.success(request, 'Visita criada com sucesso!')
                return redirect('agenda:visit_list')
            except Exception as e:
                messages.error(request, f'Erro ao salvar visita: {str(e)}')
        else:
            messages.error(request, 'Erro ao criar visita. Por favor, corrija os erros abaixo.')
    else:
        # Set initial time to next hour by default, in local time
        next_hour = timezone.localtime().replace(minute=0, second=0, microsecond=0) + timezone.timedelta(hours=1)
        form = VisitForm(initial={'scheduled_date': next_hour})
    return render(request, 'agenda/visits/form.html', {'form': form})

@login_required
def visit_edit(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    if request.method == 'POST':
        form = VisitForm(request.POST, instance=visit)
        if form.is_valid():
            try:
                visit = form.save(commit=False)
                # Ensure the datetime is timezone-aware
                if visit.scheduled_date:
                    if not timezone.is_aware(visit.scheduled_date):
                        visit.scheduled_date = timezone.make_aware(visit.scheduled_date)
                    # Ensure the timezone is set to the current timezone
                    visit.scheduled_date = timezone.localtime(visit.scheduled_date)
                
                visit.save()
                # Clear all existing messages
                storage = messages.get_messages(request)
                storage.used = True
                # Add success message
                messages.success(request, 'Visita atualizada com sucesso!')
                return redirect('agenda:visit_detail', pk=visit.pk)
            except Exception as e:
                messages.error(request, f'Erro ao atualizar visita: {str(e)}')
    else:
        # Ensure we're working with a timezone-aware datetime
        if visit.scheduled_date and not timezone.is_aware(visit.scheduled_date):
            visit.scheduled_date = timezone.make_aware(visit.scheduled_date)
        form = VisitForm(instance=visit)
    return render(request, 'agenda/visits/form.html', {'form': form, 'visit': visit})

@login_required
def visit_delete(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    if request.method == 'POST':
        visit.delete()
        # Clear all existing messages
        storage = messages.get_messages(request)
        storage.used = True
        # Add success message
        messages.success(request, 'Visita excluída com sucesso!')
        return redirect('agenda:visit_list')
    return render(request, 'agenda/visits/confirm_delete.html', {'visit': visit})

@login_required
def export_visits_to_excel(request, visits, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Visitas"
    
    # Headers - using fields that exist in your model
    headers = [
        'Título', 'Nome', 'Email', 'Telefone', 'Morada', 
        'Tipo de Visita', 'Status', 'Data/Hora Agendada', 
        'Preço', 'Tipo de Transação', 'Comentários', 'Criado por'
    ]
    
    # Style header
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Add data
    for row_num, visit in enumerate(visits, 2):
        ws.cell(row=row_num, column=1, value=visit.title or '')
        ws.cell(row=row_num, column=2, value=visit.name or '')
        ws.cell(row=row_num, column=3, value=visit.email or '')
        ws.cell(row=row_num, column=4, value=visit.phone or '')
        ws.cell(row=row_num, column=5, value=visit.address or '')
        ws.cell(row=row_num, column=6, value=visit.get_visit_type_display() if hasattr(visit, 'get_visit_type_display') else '')
        ws.cell(row=row_num, column=7, value=visit.get_status_display() if hasattr(visit, 'get_status_display') else '')
        ws.cell(row=row_num, column=8, value=visit.scheduled_date.strftime("%d/%m/%Y %H:%M") if visit.scheduled_date else '')
        ws.cell(row=row_num, column=9, value=visit.price or '')
        ws.cell(row=row_num, column=10, value=visit.get_transaction_type_display() if hasattr(visit, 'get_transaction_type_display') else '')
        ws.cell(row=row_num, column=11, value=visit.comment or '')  # Fixed: using comment (singular)
        ws.cell(row=row_num, column=12, value=getattr(visit.created_by, 'username', '') if visit.created_by else '')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value or '')) > max_length:
                    max_length = len(str(cell.value or ''))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
    
# User management views

def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('agenda:dashboard')
        else:
            messages.error(request, 'Nome de usuário ou senha inválidos.')
    return render(request, 'agenda/auth/login.html')

def user_logout(request):
    logout(request)
    return redirect('agenda:login')

class CustomPasswordResetView(PasswordResetView):
    template_name = 'password_reset_form.html'
    email_template_name = 'password_reset_email.html'
    subject_template_name = 'password_reset_subject.txt'
    success_url = '/login/'
